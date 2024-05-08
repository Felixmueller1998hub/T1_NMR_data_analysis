# -*- coding: utf-8 -*-
"""
Created on Mon Feb 12 13:15:26 2024
Gets FWHM for deconvoluted data and writes it into an excel sheet
@author: felix
"""
#%%
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from lmfit.models import LorentzianModel
from lmfit.models import ConstantModel
from lmfit import Model
import sys
import openpyxl


 
#%% functions

def lorentzian(x,center,A,sigma):
    return (A/np.pi)*(sigma / ((x - center)**2 + sigma**2))

def mixed(x, center, A, sigma, alpha):
    lorentzian_part = (A / np.pi) * (sigma / ((x - center) ** 2 + sigma ** 2))
    gaussian_part = (A / (sigma*np.sqrt(2*np.pi)))* np.exp(-0.5 * ((x - center) / sigma) ** 2)
    return (1 - alpha)*lorentzian_part + alpha * gaussian_part

def background(x, lower_limit, upper_limit):
    # Select data points outside the specified limits
    background_data = measurements.loc[(measurements['chemical shift'] < lower_limit) | (measurements['chemical shift'] > upper_limit)]
    # Calculate the average intensity of the selected data
    avg_intensity = background_data['intensity'].mean()
    # Create an array with the same length as x, filled with the average intensity
    background_values = np.full_like(x, avg_intensity)
    return background_values
        
def breakfunction():
    while True:
        # Prompt the user for input
        user_input = input("Please enter 'y' if deconvolution is right and 'n' to stop (y/n): ")
        # Check user input
        if user_input.lower() == 'y':
            print("Deconvolution is right. Continuing with the run.")
            return True  # Continue with the rest of the code
        elif user_input.lower() == 'n':
            print("Stopping the run.")
            sys.exit()  # Stop the entire Python script
        else:
            print("Invalid input. Please enter 'y' or 'n'.")
            # Ask again  
 
def import_NMR(filename_measurement, start_point, end_point, lower_limit = -500, upper_limit = 500):
   
    # import the data
    measurements = pd.read_csv(filename_measurement,sep=" ", skiprows=3, header=None)
    
    # rename the column
    measurements.columns = ['intensity']
   
    # generate the chemical shift column form the start point and number of datapoints.
    measurements['chemical shift'] = np.linspace(start_point,end_point,len(measurements))
 
    # optional: only consider part of the data between the lower and upper limit
    ppm_low = measurements['chemical shift'] >= lower_limit-0.001
    ppm_high = measurements['chemical shift'] <= upper_limit+0.001
    measurements = measurements.loc[ppm_low & ppm_high]
    measurements = measurements.reset_index(drop=True)
    return measurements

#%% input for data selection and deconvolution

data_set = '18.1wArgyro_dry'
excel_path = r'C:\Users\felix\OneDrive\Desktop\Master Thesis\PEO_LiTSFI_18.1wArgyrodite.xlsx'

#title
temp = -50
filetitle= rf'18.1wArgyrodry{temp}C'

#Peak centers, start wide and then go narrow

PEO_center = -2
PEO_plmn =  10         #final plus-minus as small as possible
PEO_sig = 20
sigpm = 10
Arg_center = -0
Arg_plmn = 5                #final plus-minus as small as possible

#%%
################################### Adjust parameters as you see fit ########################################################### 
def fit_Li6_LITFSI_with_Argyrodite(measurements, plot = False):
    # Model for PEO Peak
    mod_PEO = Model(mixed)
    pars = mod_PEO.guess(measurements['intensity'], x=measurements['chemical shift'])
    pars['alpha'].set(value=1, min = 0, max = 1)
    pars['amplitude'].set(value=10**8, min = 10**4, max = 10**10)
    pars['center'].set(value= PEO_center, min = PEO_center - PEO_plmn, max = PEO_center + PEO_plmn)
    pars['sigma'].set(value = PEO_sig, min = PEO_sig - sigpm, max = PEO_sig + sigpm) 
    
    # Model for Argyrodite
    mod_Argyrodite = LorentzianModel(prefix = 'Argyrodite_')
    pars.update(mod_Argyrodite.make_params())
    pars['Argyrodite_amplitude'].set(value=10**10, min = 10**3, max = 10**11)
    pars['Argyrodite_center'].set(value= Arg_center, min = Arg_center - Arg_plmn, max = Arg_center + Arg_plmn)
    pars['Argyrodite_sigma'].set(value=1, min = 0.3, max = 20)
    
   # Calculate background using the custom function
    background_values = background(measurements['chemical shift'], lower_limit, upper_limit) 
   
    # Model for constant background
    mod_background = ConstantModel(prefix='background_')
    pars.update(mod_background.make_params())
    pars['background_c'].set(value=background_values.mean())
    
    # add all the diferent models of the different contributions to one fit function
    mod = mod_PEO + mod_background + mod_Argyrodite
   
    # perform the fit
    out = mod.fit(measurements['intensity'], pars, x=measurements['chemical shift'])
    
    # Accessing standard errors of parameters
    std_errors = np.array([out.params[param].stderr for param in out.params])
    
    # perform the fit
    out = mod.fit(measurements['intensity'], pars, x=measurements['chemical shift'])
    if plot == True:
        print(out.fit_report(min_correl=0.25))
        plt.figure()
        out.plot_fit()
        
    return out.best_fit, out.best_values, out.fit_report(show_correl=False), std_errors
##############################################################################################################################

#%%
################################## Import the data ###############################################
filename_measurement = rf"C:\Users\felix\NMRdeconvoluted_txt\{data_set}\{filetitle}.txt"


with open(filename_measurement, 'r') as file:
     start_point = float(file.readline().strip().split(": ")[1])
     end_point = float(file.readline().strip().split(": ")[1])

#import the data
measurements = import_NMR(filename_measurement, start_point, end_point, lower_limit=-400, upper_limit=400)
# Limits
lower_limit = -60
upper_limit = 60

# Fit the Data
fit, fitted_params, report, std_errors = fit_Li6_LITFSI_with_Argyrodite(measurements, plot=False)
wlarmor = 116.642
# Calculate FWHM and Maximum Amplitude
fwhm_PEO = 2+(fitted_params['alpha']*0.3548)*fitted_params['sigma']
fwhm_Argyrodite = 2*fitted_params['Argyrodite_sigma']
height_PEO = fitted_params['amplitude']/(fitted_params['sigma']*np.pi)
height_Argyrodite = fitted_params['Argyrodite_amplitude']/(fitted_params['Argyrodite_sigma']*np.pi)
center_PEO = fitted_params['center']
center_Argyrodite = fitted_params['Argyrodite_center']
error_fwhmPEO = 2*std_errors[2]
error_fwhmArg = 2*std_errors[7]

print(report)

# Plot the data and fit in one figure
plt.figure(figsize = (10, 10/1.62), dpi= 130)
plt.plot(measurements['chemical shift'], measurements['intensity'],label = 'data', linewidth = 0.6)
plt.plot(measurements['chemical shift'], fit,'-',label = 'fit')
plt.plot(measurements['chemical shift'], -10000+lorentzian(measurements['chemical shift'], fitted_params['center'], fitted_params['amplitude'], fitted_params['sigma']),'--',label = 'PEO')
plt.plot(measurements['chemical shift'], -10000+lorentzian(measurements['chemical shift'], fitted_params['Argyrodite_center'], fitted_params['Argyrodite_amplitude'], fitted_params['Argyrodite_sigma']),'--',label = 'Argyrodite')
# Flip the x-axis and put limits
plt.xlim(lower_limit, upper_limit)
plt.gca().invert_xaxis()
plt.legend(frameon=False,fontsize = 12)     
plt.xlabel('Chemical shift [ppm]',fontsize = 16)
plt.ylabel('Intensity',fontsize = 16)
plt.yticks(fontsize=14)
plt.xticks(fontsize=14)
plt.tight_layout()
plt.title('$PEO_{18}LiTFSI$ /10wt.% $Li_6PS_5Cl$ at -40°C')
plt.show()

#breakfunction to confirm if everything is right
breakfunction()
    
#%%
########################################### write FWHM into excel file ###############################################

#write into this excel sheet in the first sheet with the name LithiumFWHM
#and the cloumn FHWM Argyrodite (Hz) and FHWM PEO (Hz)
# Load the existing Excel file
book = openpyxl.load_workbook(excel_path)

# Access the sheet named 'LithiumFWHM' (or create it if it doesn't exist)
sheet_name = 'LithiumFWHM'
if sheet_name in book.sheetnames:
    sheet = book[sheet_name]
else:
    sheet = book.create_sheet(title=sheet_name)

# Define the column names
column_names = ['Temperature (°C)', 'FWHM PEO (ppm)', ]


# Write the values to the Excel sheet
data_row = [temp, fwhm_PEO, error_fwhmPEO]
sheet.append(data_row)

# Save the changes to the Excel file
book.save(excel_path)

print(data_row)
